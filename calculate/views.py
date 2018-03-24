# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.http import HttpResponse

from django.shortcuts import render
from django.http import JsonResponse
from django.shortcuts import redirect

# import for working with excel
import numbers
import abc
import logging
import openpyxl as xl
import pyxl_utils as xlu
from openpyxl.formula import Tokenizer

from .models import Sheet




# Create your views here.
def generate_sheets():
    sheet = Sheet.objects.get(id=1)
    sheet_parameters = sheet.get_sheet_parameters()
    print (sheet_parameters)
    html_input_string_sheet1 = ""
    for i in range(1, 81):
        if "A%s" % i in sheet_parameters:
            html_input_string_sheet1 += "<input type='text' id='A%s' name='input_A%s' placeholder='%s' data-formula='%s'>" % (i, i, sheet_parameters["A%s" % i], sheet_parameters["A%s" % i])
        else:
            html_input_string_sheet1 += "<input type='text' id='A%s' name='input_A%s'  placeholder='A%s'>" % (i, i, i)

    html_input_string_sheet2 = ""
    for i in range(1, 81):
        if "B%s" % i in sheet_parameters:
            html_input_string_sheet2 += "<input type='text' id='B%s' name='input_B%s' placeholder='%s' data-formula='%s'>" % (i, i, sheet_parameters["B%s" % i], sheet_parameters["B%s" % i])
        else:
            html_input_string_sheet2 += "<input type='text' id='B%s' name='input_B%s'  placeholder='B%s'>" % (i, i, i)

    html_input_string_sheet3 = ""
    for i in range(1, 81):
        if "C%s" % i in sheet_parameters:
            html_input_string_sheet3 += "<input type='text' id='C%s' name='input_C%s' placeholder='%s' data-formula='%s'>" % (i, i, sheet_parameters["C%s" % i], sheet_parameters["C%s" % i])
        else:
            html_input_string_sheet3 += "<input type='text' id='C%s' name='input_C%s'  placeholder='C%s'>" % (i, i, i)

    sheets = {}
    sheets['sheet1_inputs'] = html_input_string_sheet1
    sheets['sheet2_inputs'] = html_input_string_sheet2
    sheets['sheet3_inputs'] = html_input_string_sheet3
    return sheets


def index(request):
    context = generate_sheets()
    return render(request, "index.html", context)


def admin(request):
    if request.method == "GET":
        context = generate_sheets()
        context['is_admin'] = True
        return render(request, "index.html", context)
    else:
        formula_dict = {}
        for key in request.POST:
            if key.startswith('input'):
                if request.POST[key]:
                    formula_dict[key[6:]] = request.POST[key]
        sheet = Sheet.objects.get(id=1)
        sheet.set_sheet_parameters(formula_dict)
        return redirect('/calculation/admin')




def ajax_req(request):
    infile = 'forTest.xlsx'
    cell_range = 'A6:A7'

    wb = xl.load_workbook(infile)
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])  # or use wb.active

    # read from json
    ws['A1'] = int(request.POST.get("A1", 0))
    ws['A2'] = int(request.POST.get("A2", 0))
    ws['A3'] = int(request.POST.get("A3", 0))
    ws['A4'] = int(request.POST.get("A4", 0))
    ws['A5'] = int(request.POST.get("A5", 0))
    ws['A6'] = int(request.POST.get("A6", 0))
    json_return = {}
    json_list = []
    for row in ws.iter_rows(cell_range):
        for cell in row:
            if cell.data_type == 'f':
                try:
                    value = eval_formula(ws, cell.value)
                    print('Value returned from eval_formula ({}): {}'.format(cell.value, value))
                    json_list.append('Value returned from eval_formula ({}): {}'.format(cell.value, value))
                except KeyError as e:
                    print('Unknown formula or operand... ignoring cell (%s)' % cell.value)  # ignore unknown formula
            else:
                print(cell.value)
    json_return['values'] = json_list
    print(json_return)
    return JsonResponse(json_return)


def get_cell_value(ws, cell, none_as_zero=True):
    """
    Returns: the cell value. If the cell contains a formula, the formula value is evaluated and returned
    """
    if cell.value is None:
        return 0 if none_as_zero else None
    if xlu.is_formula(cell):
        return eval_formula(ws, cell.value)
    else:
        return cell.value


def get_cell(ws, s, none_as_zero=True):
    """
    ws - openpyxl worksheet
    s - string for the cell location (e.g. 'C1') or range (e.g. 'B1:C12')
    Return the value for the cell. If s is a range, returns a tuple of the cell values.
    """
    if s.find(':') == -1:  # simple cell
        return get_cell_value(ws, ws[s], none_as_zero)
    else:  # it's a range
        skip_nones = not none_as_zero  # don't include None values in the tuple if none_as_zero is False
        return tuple(get_cell_value(ws, c, none_as_zero) for c in xlu.cell_range_generator(ws, s, skip_nones))


infix_swap = {'^': '**', '=': '==', '<>': '!=', '&': '+'}
func_swap = {'ABS(': 'abs(', 'SUM(': 'sum(', 'MIN(': 'min(', 'MAX(': 'max(', 'AVERAGE(': 'average(', 'IF(': 'excel_if(',
             'AND(': 'excel_and(', 'OR(': 'excel_or('}
operand_swap = {'TRUE': 'True', 'FALSE': 'False'}


def average(*args):
    # Emulate the Excel average formula
    total = items = 0

    for arg in args:
        if isinstance(arg, numbers.Number):
            total += arg
            items += 1
        elif isinstance(arg, collections.abc.Iterable):
            total += sum(arg)
            items += len(arg)
        else:
            s = 'Average function cannot handle argument type ({})'.format(arg)
            logging.error(s)
            raise RuntimeError(s)

    return total/items


def excel_if(test, yes, no):
    # Emulate the Excel if statement
    return yes if test else no


def excel_and(*args):
    # Emulate the Excel and statement
    for arg in args:
        if not arg:
            return False
    return True


def excel_or(*args):
    # Emulate the Excel or statement
    for arg in args:
        if arg:
            return True
    return False


def eval_formula(ws, formula_str):
    eval_str = ''
    tok = Tokenizer(formula_str)
    #tok.parse()
    #print("\n".join("%12s%11s%9s" % (t.value, t.type, t.subtype) for t in tok.items))

    for t in tok.items:
        if t.subtype == 'RANGE':
            range_str = 'get_cell(ws,"{}")'.format(t.value)
            eval_str += range_str
        elif t.type == 'OPERATOR-INFIX':
            try:
                eval_str += infix_swap[t.value]
            except KeyError:
                eval_str += t.value
        elif t.type == 'OPERAND':
            try:
                eval_str += operand_swap[t.value]
            except KeyError:
                eval_str += t.value
        elif t.type == 'FUNC' and t.subtype == 'OPEN':
            try:
                eval_str += func_swap[t.value]
            except KeyError:
                logging.error('Unsupported function in formula ({})'.format(t.value[:-1]))
                raise
        else:
            eval_str += t.value

    #print('formula_str=%s   eval_str=%s' % (formula_str, eval_str))
    # value = eval(eval_str)
    # return value
    if len(eval_str) == 0:
        pass

    try:
        return eval(eval_str)
    except:
        logging.error('Exception evaluating eval_str=\'{}\''.format(eval_str))
        raise